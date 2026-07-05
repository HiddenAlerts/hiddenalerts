"""Read-only export of the current V1 review queue (publish_decision = 'review').

EXPORT ONLY — runs SELECT-only queries via the ORM and NEVER commits or mutates.
No backfill/apply/migration/deploy. Produces, under ``backend/reports/``:
  - v1_review_queue_export_<ts>.csv
  - v1_review_queue_export_<ts>.xlsx   (if openpyxl is available)
  - v1_review_queue_export_summary_<ts>.md

Safe to share with Ken: only alert data + blank ``ken_decision`` / ``ken_notes``
columns for him to fill. No secrets, credentials, or environment values.

Run (read-only, against the prod DB the app is configured for):
    docker compose -f docker-compose.yml run --rm --no-deps -v "$(pwd)":/app app \
      python scripts/export_v1_review_queue.py
"""
from __future__ import annotations

import asyncio
import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import case, select
from sqlalchemy.orm import selectinload

from app.database import AsyncSessionLocal
from app.models.processed_alert import ProcessedAlert
from app.models.raw_item import RawItem

OUT_DIR = Path("reports")

COLUMNS = [
    "alert_id", "title", "source_name", "item_url",
    "source_published_at", "processed_at", "primary_category",
    "risk_level", "risk_band",
    "signal_score_total", "signal_score_internal", "signal_score_100",
    "publish_decision", "publish_decision_reason", "pending_review_reason",
    "publication_state_source", "publication_state_updated_at",
    "summary", "matched_keywords", "entities", "source_credibility",
    "score_source_credibility", "score_financial_impact", "score_victim_scale",
    "score_cross_source", "score_trend_acceleration",
    "suggested_review_action", "ken_decision", "ken_notes",
]

_BAND_ORDER = case(
    (ProcessedAlert.risk_band == "critical", 1),
    (ProcessedAlert.risk_band == "high", 2),
    (ProcessedAlert.risk_band == "medium", 3),
    (ProcessedAlert.risk_band == "below_60", 4),
    else_=5,
)


def _clean(value) -> str:
    """Collapse all whitespace/newlines so the cell stays on one spreadsheet line."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _iso(dt) -> str:
    return dt.isoformat() if dt is not None else ""


def _to_100(internal):
    return round(internal / 25 * 100) if internal is not None else None


def _keywords(value) -> str:
    if not value:
        return ""
    if isinstance(value, dict):
        return ", ".join(str(k) for k in value.keys())
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(x) for x in value if x)
    return _clean(value)


def _entities(entities_json) -> str:
    if not entities_json:
        return ""
    if isinstance(entities_json, dict):
        names = entities_json.get("names")
        if isinstance(names, list):
            return ", ".join(str(x) for x in names if x)
    if isinstance(entities_json, list):
        return ", ".join(str(x) for x in entities_json if x)
    return ""


def _suggest(pending_reason, risk_band) -> str:
    if pending_reason == "manual_review_only" and risk_band in ("critical", "high"):
        return "Review for possible publish"
    if pending_reason == "blocked_by_topic_scope":
        return "Likely reject / out of scope"
    if pending_reason == "blocked_by_score":
        return "Review if still relevant"
    if pending_reason == "blocked_by_source_rule":
        return "Review source-specific context"
    return "Review"


def _row(alert: ProcessedAlert) -> dict:
    raw = alert.raw_item
    source = raw.source if raw else None
    internal = alert.signal_score_total
    s100 = _to_100(internal)
    return {
        "alert_id": alert.id,
        "title": _clean(raw.title if raw else None),
        "source_name": (source.name if source else "") or "",
        "item_url": (raw.item_url if raw else "") or "",
        "source_published_at": _iso(raw.published_at if raw else None),
        "processed_at": _iso(alert.processed_at),
        "primary_category": alert.primary_category or "",
        "risk_level": alert.risk_level or "",
        "risk_band": alert.risk_band or "",
        "signal_score_total": s100,        # user-facing 0-100 (DB stores 5-25 internally)
        "signal_score_internal": internal,  # raw 5-25
        "signal_score_100": s100,           # 0-100
        "publish_decision": alert.publish_decision or "",
        "publish_decision_reason": alert.publish_decision_reason or "",
        "pending_review_reason": alert.pending_review_reason or "",
        "publication_state_source": alert.publication_state_source or "",
        "publication_state_updated_at": _iso(alert.publication_state_updated_at),
        "summary": _clean(alert.summary),
        "matched_keywords": _keywords(alert.matched_keywords),
        "entities": _entities(alert.entities_json),
        "source_credibility": source.credibility_score if source else None,
        "score_source_credibility": alert.score_source_credibility,
        "score_financial_impact": alert.score_financial_impact,
        "score_victim_scale": alert.score_victim_scale,
        "score_cross_source": alert.score_cross_source,
        "score_trend_acceleration": alert.score_trend_acceleration,
        "suggested_review_action": _suggest(alert.pending_review_reason, alert.risk_band),
        "ken_decision": "",
        "ken_notes": "",
    }


async def _fetch_rows() -> list[dict]:
    async with AsyncSessionLocal() as session:  # SELECT-only; never commits
        stmt = (
            select(ProcessedAlert)
            .join(RawItem, RawItem.id == ProcessedAlert.raw_item_id)
            .where(ProcessedAlert.publish_decision == "review")
            .options(selectinload(ProcessedAlert.raw_item).selectinload(RawItem.source))
            .order_by(
                _BAND_ORDER,
                ProcessedAlert.pending_review_reason,
                ProcessedAlert.publication_state_source,
                RawItem.published_at.desc().nullslast(),
                ProcessedAlert.processed_at.desc().nullslast(),
            )
        )
        alerts = (await session.execute(stmt)).scalars().all()
    return [_row(a) for a in alerts]


def _is_high_priority(r: dict) -> bool:
    return r["risk_band"] in ("critical", "high") and (
        r["pending_review_reason"] == "manual_review_only"
        or r["publish_decision_reason"] == "historical_reclassified_publishable"
    )


def _write_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def _counter(rows, key):
    return dict(Counter((r[key] or "(none)") for r in rows).most_common())


def _write_xlsx(rows: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def _sheet(title, data_rows, columns=COLUMNS):
        ws = wb.create_sheet(title=title)
        ws.append(columns)
        for c in range(1, len(columns) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for r in data_rows:
            ws.append([r.get(col, "") for col in columns])
        ws.freeze_panes = "A2"
        # Simple width fit (capped so summary/title cells don't blow out).
        for idx, col in enumerate(columns, start=1):
            longest = max([len(str(col))] + [len(str(r.get(col, ""))) for r in data_rows] or [0])
            ws.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 10), 60)
        return ws

    wb.remove(wb.active)  # drop default empty sheet
    _sheet("Review Queue", rows)
    _sheet("High Priority", [r for r in rows if _is_high_priority(r)])
    _sheet("Topic Scope Blocked",
           [r for r in rows if r["pending_review_reason"] == "blocked_by_topic_scope"])

    # Summary sheet — counts only.
    ws = wb.create_sheet(title="Summary")
    ws["A1"] = "V1 Review Queue — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total review-queue rows", len(rows)])
    for label, key in (("By risk_band", "risk_band"),
                       ("By pending_review_reason", "pending_review_reason"),
                       ("By publication_state_source", "publication_state_source"),
                       ("By primary_category", "primary_category"),
                       ("By source", "source_name")):
        ws.append([])
        ws.append([label])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for k, v in _counter(rows, key).items():
            ws.append([k, v])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12

    wb.save(path)


def _write_summary_md(rows: list[dict], ts_iso: str, path: Path) -> None:
    def block(title, key):
        lines = [f"**{title}**", ""]
        for k, v in _counter(rows, key).items():
            lines.append(f"- {k}: {v}")
        return "\n".join(lines) + "\n"

    top = sorted(
        rows,
        key=lambda r: ({"critical": 1, "high": 2, "medium": 3, "below_60": 4}.get(r["risk_band"], 5),
                       -(r["signal_score_100"] or 0)),
    )[:10]
    top_lines = ["| alert_id | risk_band | source | title | suggested_review_action |",
                 "|---|---|---|---|---|"]
    for r in top:
        title = (r["title"] or "")[:70].replace("|", "\\|")
        top_lines.append(
            f"| {r['alert_id']} | {r['risk_band']} | {r['source_name']} | {title} | {r['suggested_review_action']} |"
        )

    md = f"""# V1 Review Queue — Export Summary

- **Export timestamp:** {ts_iso}
- **Database / environment checked:** production (the database the app container is configured for)
- **Total review-queue rows exported:** {len(rows)}

## Counts

{block("By risk_band", "risk_band")}
{block("By pending_review_reason", "pending_review_reason")}
{block("By publication_state_source", "publication_state_source")}
{block("By primary_category", "primary_category")}

## Top 10 highest-priority rows

{chr(10).join(top_lines)}

## Notes
- **Read-only export** — no database changes were made.
- Use the **`ken_decision`** and **`ken_notes`** columns to record your review.
- Suggested actions are conservative hints only — final decisions are yours.
- **Do not use the old Jinja dashboard's approve/reject buttons for V1 review** (they predate V1 and do not write the new fields). Approved/rejected/held decisions from this sheet will be applied later through the correct V1 backend workflow.
"""
    path.write_text(md, encoding="utf-8")


async def main() -> int:
    rows = await _fetch_rows()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.now(timezone.utc)
    ts = now.strftime("%Y%m%d_%H%M%S")

    csv_path = OUT_DIR / f"v1_review_queue_export_{ts}.csv"
    md_path = OUT_DIR / f"v1_review_queue_export_summary_{ts}.md"
    _write_csv(rows, csv_path)
    _write_summary_md(rows, now.isoformat(), md_path)
    written = [str(csv_path), str(md_path)]

    xlsx_path = OUT_DIR / f"v1_review_queue_export_{ts}.xlsx"
    try:
        _write_xlsx(rows, xlsx_path)
        written.append(str(xlsx_path))
    except ImportError:
        pass  # CSV + MD already cover the requirement

    # Machine-readable result for validation (printed, never written to a shared file).
    print(json.dumps({
        "row_count": len(rows),
        "files": written,
        "all_review": all(r["publish_decision"] == "review" for r in rows),
        "no_blank_alert_id": all(r["alert_id"] for r in rows),
        "ken_cols_blank": all(r["ken_decision"] == "" and r["ken_notes"] == "" for r in rows),
        "by_risk_band": _counter(rows, "risk_band"),
        "by_pending_review_reason": _counter(rows, "pending_review_reason"),
        "by_publication_state_source": _counter(rows, "publication_state_source"),
        "by_primary_category": _counter(rows, "primary_category"),
        "top10": [
            {"alert_id": r["alert_id"], "risk_band": r["risk_band"], "source": r["source_name"],
             "title": (r["title"] or "")[:70], "suggested": r["suggested_review_action"]}
            for r in sorted(rows, key=lambda r: ({"critical": 1, "high": 2, "medium": 3, "below_60": 4}.get(r["risk_band"], 5), -(r["signal_score_100"] or 0)))[:10]
        ],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
