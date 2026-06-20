"""V1 Candidate Backfill — DRY-RUN / audit tool (Slice 7B).

Read-only validation of Ken's reviewed Candidate Alert decision file against the
database, BEFORE any apply step (Slice 7C) is built. This tool NEVER writes to
the database: it opens a session, runs SELECT-only queries, builds ORM-free
preview dicts, and emits a JSON (and optional CSV) report.

Run as a module from the backend directory:

    python -m app.tools.v1_candidate_backfill_dry_run \
      --input path/to/reviewed_candidate_alerts.csv \
      --output reports/v1_candidate_backfill_dry_run.json

Input: ``.csv`` always; ``.xlsx`` only when the optional ``openpyxl`` package is
installed (a clear error is raised otherwise — convert to CSV or install it).

The reviewed Candidate file is expected to carry 48 approved / 44 rejected /
27 keep-later decisions (total 119). Count mismatch fails by default unless
``--allow-count-mismatch`` is passed. Nothing is ever applied.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.processed_alert import ProcessedAlert
from app.models.raw_item import RawItem
from app.pipeline.publishing.constants import (
    DecisionSource,
    PendingReviewReason,
    PublishDecisionValue,
)
from app.pipeline.publishing.publishing_policy import DEFAULT_V1_POLICY
from app.pipeline.publishing.risk_bands import compute_risk_band

# Expected reviewed-file decision counts (Ken, 2026-06-11).
EXPECTED_COUNTS: dict[str, int] = {"approved": 48, "rejected": 44, "keep_later": 27}

# ---------------------------------------------------------------------------
# Column + decision vocabularies (tolerant, case-insensitive)
# ---------------------------------------------------------------------------

_ID_COLUMN_ALIASES = {
    "alert_id", "id", "processed_alert_id", "candidate_alert_id", "alert id",
}
_DECISION_COLUMN_ALIASES = {
    "decision", "ken_decision", "review_decision", "status", "review_status",
    "action", "ken decision",
    # Ken's reviewed Candidate file uses this exact header.
    "manual review decision", "manual_review_decision",
}

_APPROVED_ALIASES = {
    "approved", "approve", "publish", "backfill", "controlled backfill",
    "approved for controlled backfill", "yes",
    # Ken's file value.
    "approve for controlled backfill",
}
_REJECTED_ALIASES = {
    "rejected", "reject", "exclude", "excluded", "false_positive",
    "false positive", "no",
}
_KEEP_LATER_ALIASES = {
    "keep_later", "keep later", "keep", "hold", "manual_hold", "manual hold",
    "future review", "review later",
    # Ken's file value.
    "keep for later",
}

# Per-decision preview of the V1 state a future apply step (Slice 7C) WOULD set.
# Uses canonical enum values + DEFAULT_V1_POLICY.version; never written here. The
# alert-specific ``risk_band`` is added per-row in validate_db().
_WOULD_CHANGE = {
    "approved": {
        "is_published": True,
        "publish_decision": PublishDecisionValue.AUTO_PUBLISH.value,
        "publish_decision_reason": "candidate_backfill_approved",
        "pending_review_reason": None,
        "is_excluded": False,
        "excluded_reason": None,
        "is_manual_hold": False,
        "published_by_rule": False,
        "publication_state_source": DecisionSource.CANDIDATE_BACKFILL.value,
        "publishing_policy_version": DEFAULT_V1_POLICY.version,
    },
    "rejected": {
        "is_published": False,
        "publish_decision": PublishDecisionValue.EXCLUDE.value,
        "publish_decision_reason": "candidate_backfill_rejected",
        "pending_review_reason": PendingReviewReason.MANUAL_REJECTED.value,
        "is_excluded": True,
        "excluded_reason": "candidate_backfill_rejected",
        "is_manual_hold": False,
        "published_by_rule": False,
        "publication_state_source": DecisionSource.CANDIDATE_BACKFILL.value,
        "publishing_policy_version": DEFAULT_V1_POLICY.version,
    },
    "keep_later": {
        "is_published": False,
        "publish_decision": PublishDecisionValue.HOLD.value,
        "publish_decision_reason": "candidate_backfill_keep_later",
        "pending_review_reason": PendingReviewReason.MANUAL_HOLD.value,
        "is_excluded": False,
        "excluded_reason": None,
        "is_manual_hold": True,
        "published_by_rule": False,
        "publication_state_source": DecisionSource.CANDIDATE_BACKFILL.value,
        "publishing_policy_version": DEFAULT_V1_POLICY.version,
    },
}


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


@dataclass
class ParsedFile:
    id_column: str | None = None
    decision_column: str | None = None
    counts: dict[str, int] = field(default_factory=lambda: {"approved": 0, "rejected": 0, "keep_later": 0})
    valid: dict[int, str] = field(default_factory=dict)  # alert_id -> decision (deduped)
    invalid_rows: list[dict] = field(default_factory=list)
    invalid_decisions: list[dict] = field(default_factory=list)
    duplicate_same: list[int] = field(default_factory=list)
    duplicate_conflicting: list[int] = field(default_factory=list)
    total_rows: int = 0


def normalize_decision(value: Any) -> str | None:
    """Map a raw decision cell to 'approved' | 'rejected' | 'keep_later' | None."""
    if value is None:
        return None
    token = str(value).strip().lower()
    if not token:
        return None
    if token in _APPROVED_ALIASES:
        return "approved"
    if token in _REJECTED_ALIASES:
        return "rejected"
    if token in _KEEP_LATER_ALIASES:
        return "keep_later"
    return None


def _to_int_id(value: Any) -> int | None:
    """Coerce an id cell (int / '123' / 123.0 / '123.0') to int, else None."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        try:
            f = float(s)
        except ValueError:
            return None
        return int(f) if f.is_integer() else None


def detect_columns(headers: list[str]) -> tuple[str | None, str | None]:
    """Return (id_column, decision_column) original-cased, via alias matching."""
    id_col = decision_col = None
    for h in headers:
        key = (h or "").strip().lower()
        if id_col is None and key in _ID_COLUMN_ALIASES:
            id_col = h
        if decision_col is None and key in _DECISION_COLUMN_ALIASES:
            decision_col = h
    return id_col, decision_col


def parse_records(headers: list[str], rows: list[dict]) -> ParsedFile:
    """Parse already-read header/row dicts into a ParsedFile (no DB)."""
    out = ParsedFile(total_rows=len(rows))
    out.id_column, out.decision_column = detect_columns(headers)
    if out.id_column is None or out.decision_column is None:
        # Can't proceed without both columns — every row is invalid context.
        out.invalid_rows.append(
            {"reason": "missing_required_columns",
             "id_column": out.id_column, "decision_column": out.decision_column,
             "headers": headers}
        )
        return out

    # alert_id -> set of normalized decisions seen
    seen: dict[int, set[str]] = {}
    for idx, row in enumerate(rows):
        raw_id = row.get(out.id_column)
        raw_decision = row.get(out.decision_column)
        alert_id = _to_int_id(raw_id)
        if alert_id is None:
            out.invalid_rows.append({"row": idx + 2, "raw_id": raw_id, "reason": "missing_or_non_integer_id"})
            continue
        decision = normalize_decision(raw_decision)
        if decision is None:
            out.invalid_decisions.append({"row": idx + 2, "alert_id": alert_id, "raw_decision": raw_decision})
            continue
        seen.setdefault(alert_id, set()).add(decision)

    for alert_id, decisions in seen.items():
        if len(decisions) > 1:
            out.duplicate_conflicting.append(alert_id)
            continue
        decision = next(iter(decisions))
        out.valid[alert_id] = decision
        out.counts[decision] += 1
        # Track ids that appeared more than once with a consistent decision.
        occurrences = sum(
            1 for r in rows if _to_int_id(r.get(out.id_column)) == alert_id
        )
        if occurrences > 1:
            out.duplicate_same.append(alert_id)

    out.duplicate_conflicting.sort()
    out.duplicate_same.sort()
    return out


def read_input(path: Path, *, sheet: str | None = None) -> tuple[list[str], list[dict]]:
    """Read a .csv or .xlsx decision file into (headers, list[row-dict])."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet=sheet)
    raise ValueError(f"Unsupported input type {suffix!r}. Use .csv or .xlsx.")


def _read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        raw_fields = reader.fieldnames or []
        headers = [h.strip() for h in raw_fields]
        # Re-key each row to the STRIPPED headers so spaced headers like
        # " Alert ID " / " Manual Review Decision " resolve correctly.
        rows = [
            {h.strip(): r.get(h) for h in raw_fields}
            for r in reader
        ]
    return headers, rows


def _sheet_to_rows(ws) -> tuple[list[str], list[dict]]:
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return [], []
    headers = [("" if h is None else str(h)).strip() for h in raw[0]]
    rows: list[dict] = []
    for r in raw[1:]:
        if all(c is None for c in r):
            continue
        rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
    return headers, rows


def _read_xlsx(path: Path, *, sheet: str | None = None) -> tuple[list[str], list[dict]]:
    try:
        import openpyxl  # noqa: PLC0415 — optional, lazily imported
    except ImportError as exc:  # pragma: no cover - env-dependent
        raise RuntimeError(
            "XLSX input requires the 'openpyxl' package, which is not installed. "
            "Convert the file to CSV (recommended) or install openpyxl in a "
            "later dependency slice."
        ) from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet!r} not found. Available: {wb.sheetnames}")
        return _sheet_to_rows(wb[sheet])
    # Multi-sheet workbooks (e.g. Ken's Read Me / Summary / Review Data): pick the
    # first sheet whose header row exposes BOTH an id and a decision column.
    fallback = None
    for sn in wb.sheetnames:
        headers, rows = _sheet_to_rows(wb[sn])
        if fallback is None:
            fallback = (headers, rows)
        id_col, dec_col = detect_columns(headers)
        if id_col is not None and dec_col is not None:
            return headers, rows
    return fallback if fallback is not None else ([], [])


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_counts(
    counts: dict[str, int], expected: dict[str, int], allow_mismatch: bool
) -> dict:
    errors = []
    for key in ("approved", "rejected", "keep_later"):
        if counts.get(key, 0) != expected.get(key, 0):
            errors.append(
                f"{key}: got {counts.get(key, 0)}, expected {expected.get(key, 0)}"
            )
    passed = not errors or allow_mismatch
    return {"passed": passed, "strict_match": not errors, "errors": errors}


async def validate_db(session: AsyncSession, parsed: ParsedFile) -> tuple[dict, dict]:
    """Read-only DB validation. Returns (db_validation, planned_actions).

    NEVER mutates: only SELECTs and builds plain dicts.
    """
    planned: dict[str, list] = {"approved": [], "rejected": [], "keep_later": []}
    missing_ids: list[int] = []
    already_published_ids: list[int] = []
    approved_irrelevant_ids: list[int] = []

    if parsed.valid:
        result = await session.execute(
            select(ProcessedAlert)
            .where(ProcessedAlert.id.in_(list(parsed.valid.keys())))
            .options(selectinload(ProcessedAlert.raw_item).selectinload(RawItem.source))
        )
        by_id = {a.id: a for a in result.scalars().all()}
    else:
        by_id = {}

    for alert_id, decision in sorted(parsed.valid.items()):
        alert = by_id.get(alert_id)
        if alert is None:
            missing_ids.append(alert_id)
            continue

        source_name = source_cred = None
        if alert.raw_item and alert.raw_item.source:
            source_name = alert.raw_item.source.name
            source_cred = alert.raw_item.source.credibility_score

        if alert.is_published:
            already_published_ids.append(alert_id)
        if decision == "approved" and not alert.is_relevant:
            approved_irrelevant_ids.append(alert_id)

        current = {
            "is_published": alert.is_published,
            "is_relevant": alert.is_relevant,
            "signal_score_total": alert.signal_score_total,
            "risk_level": alert.risk_level,
            "risk_band": alert.risk_band or compute_risk_band(alert.signal_score_total).value,
            "publish_decision": alert.publish_decision,
            "pending_review_reason": alert.pending_review_reason,
            "is_excluded": alert.is_excluded,
            "is_manual_hold": alert.is_manual_hold,
            "published_by_rule": alert.published_by_rule,
            "publication_state_source": alert.publication_state_source,
            "primary_category": alert.primary_category,
            "source_name": source_name,
            "source_credibility": source_cred,
            "processed_at": alert.processed_at.isoformat() if alert.processed_at else None,
            "published_at": alert.published_at.isoformat() if alert.published_at else None,
        }
        # Complete the preview with the alert-specific computed risk band.
        would_change = dict(_WOULD_CHANGE[decision])
        would_change["risk_band"] = compute_risk_band(alert.signal_score_total).value
        planned[decision].append(
            {"alert_id": alert_id, "current": current, "would_change": would_change}
        )

    # Blocker subsets: approved rows that are already published or irrelevant.
    approved_already_published = sorted(
        aid for aid in already_published_ids if parsed.valid.get(aid) == "approved"
    )
    db_validation = {
        "passed": not (
            missing_ids or approved_already_published or approved_irrelevant_ids
        ),
        "missing_ids": sorted(missing_ids),
        "already_published_ids": sorted(already_published_ids),
        "approved_already_published_ids": approved_already_published,
        "approved_irrelevant_ids": sorted(approved_irrelevant_ids),
        "duplicate_ids": parsed.duplicate_same,
        "conflicting_duplicate_ids": parsed.duplicate_conflicting,
    }
    return db_validation, planned


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


async def run_dry_run(
    input_path: str | Path,
    *,
    session: AsyncSession,
    expected_counts: dict[str, int] | None = None,
    allow_count_mismatch: bool = False,
    sheet: str | None = None,
) -> dict:
    """Full read-only dry run; returns the report dict. Never mutates the DB."""
    expected = expected_counts or EXPECTED_COUNTS
    path = Path(input_path)
    headers, rows = read_input(path, sheet=sheet)
    parsed = parse_records(headers, rows)

    count_validation = validate_counts(parsed.counts, expected, allow_count_mismatch)
    db_validation, planned = await validate_db(session, parsed)

    blockers: list[str] = []
    if parsed.id_column is None or parsed.decision_column is None:
        blockers.append("missing_required_columns")
    if parsed.invalid_rows:
        blockers.append(f"invalid_rows={len(parsed.invalid_rows)}")
    if parsed.invalid_decisions:
        blockers.append(f"invalid_decisions={len(parsed.invalid_decisions)}")
    if parsed.duplicate_conflicting:
        blockers.append(f"conflicting_duplicate_ids={len(parsed.duplicate_conflicting)}")
    if not count_validation["strict_match"] and not allow_count_mismatch:
        blockers.append("count_mismatch")
    if db_validation["missing_ids"]:
        blockers.append(f"missing_ids={len(db_validation['missing_ids'])}")
    if db_validation["approved_already_published_ids"]:
        blockers.append(
            f"approved_already_published={len(db_validation['approved_already_published_ids'])}"
        )
    if db_validation["approved_irrelevant_ids"]:
        blockers.append(
            f"approved_irrelevant={len(db_validation['approved_irrelevant_ids'])}"
        )

    warnings: list[str] = []
    if parsed.duplicate_same:
        warnings.append(f"duplicate_same_decision_ids={len(parsed.duplicate_same)}")
    if not count_validation["strict_match"] and allow_count_mismatch:
        warnings.append("count_mismatch_allowed")

    total_valid = len(parsed.valid)
    invalid = len(parsed.invalid_rows) + len(parsed.invalid_decisions)
    passed = not blockers

    return {
        "mode": "dry_run",
        "input_file": str(path),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "detected_columns": {"id": parsed.id_column, "decision": parsed.decision_column},
        "expected_counts": {**expected, "total": sum(expected.values())},
        "actual_counts": {
            "approved": parsed.counts["approved"],
            "rejected": parsed.counts["rejected"],
            "keep_later": parsed.counts["keep_later"],
            "invalid": invalid,
            "conflicting": len(parsed.duplicate_conflicting),
            "total_valid": total_valid,
            "total_rows": parsed.total_rows,
        },
        "count_validation": count_validation,
        "db_validation": db_validation,
        "planned_actions": planned,
        "invalid_rows": parsed.invalid_rows,
        "invalid_decisions": parsed.invalid_decisions,
        "blockers": blockers,
        "warnings": warnings,
        "passed": passed,
    }


def console_summary(report: dict) -> str:
    exp = report["expected_counts"]
    act = report["actual_counts"]
    db = report["db_validation"]
    cols = report["detected_columns"]
    return "\n".join(
        [
            "Candidate Backfill Dry Run",
            f"Detected ID column: {cols['id']!r}",
            f"Detected decision column: {cols['decision']!r}",
            f"Input rows: {act['total_rows']}",
            f"Approved: {act['approved']} / expected {exp['approved']}",
            f"Rejected: {act['rejected']} / expected {exp['rejected']}",
            f"Keep later: {act['keep_later']} / expected {exp['keep_later']}",
            f"Invalid rows/decisions: {act['invalid']}",
            f"Conflicting duplicate IDs: {len(db['conflicting_duplicate_ids'])}",
            f"Duplicate (same decision) IDs: {len(db['duplicate_ids'])}",
            f"Missing IDs: {len(db['missing_ids'])}",
            f"Already published: {len(db['already_published_ids'])}",
            f"Approved irrelevant: {len(db['approved_irrelevant_ids'])}",
            f"Blockers: {report['blockers']}",
            f"Passed: {report['passed']}",
            "No database changes were applied.",
        ]
    )


def _write_csv_rows(report: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "decision", "alert_id", "current_is_published", "current_is_relevant",
            "signal_score_total", "risk_band", "primary_category", "source_name",
            "source_credibility", "would_publish_decision", "would_publication_state_source",
        ])
        for decision, items in report["planned_actions"].items():
            for it in items:
                cur, wc = it["current"], it["would_change"]
                w.writerow([
                    decision, it["alert_id"], cur["is_published"], cur["is_relevant"],
                    cur["signal_score_total"], cur["risk_band"], cur["primary_category"],
                    cur["source_name"], cur["source_credibility"],
                    wc.get("publish_decision"), wc.get("publication_state_source"),
                ])


async def _main_async(args: argparse.Namespace) -> int:
    from app.database import AsyncSessionLocal  # local import: keep module test-friendly

    async with AsyncSessionLocal() as session:
        report = await run_dry_run(
            args.input,
            session=session,
            allow_count_mismatch=args.allow_count_mismatch,
            sheet=args.sheet,
        )

    if args.output:
        out = Path(args.output)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(report, indent=2), encoding="utf-8")
    if args.output_csv:
        _write_csv_rows(report, Path(args.output_csv))

    print(console_summary(report))
    if args.output:
        print(f"\nJSON report written to: {args.output}")
    return 0 if report["passed"] else 1


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="DRY-RUN audit of Ken's reviewed Candidate Alert decision file. "
        "Read-only — never writes to the database."
    )
    parser.add_argument("--input", required=True, help="Path to reviewed decisions (.csv or .xlsx)")
    parser.add_argument("--output", help="Path to write the JSON dry-run report")
    parser.add_argument("--output-csv", help="Optional path to write row-level planned actions")
    parser.add_argument("--sheet", help="XLSX sheet name (default: auto-detect the data sheet)")
    parser.add_argument(
        "--allow-count-mismatch", action="store_true",
        help="Do not fail when file counts differ from 48/44/27 (default: strict).",
    )
    args = parser.parse_args(argv)
    return asyncio.run(_main_async(args))


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
