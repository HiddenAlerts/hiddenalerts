"""V1 Review-Queue Decisions — DRY-RUN + guarded APPLY (Ken's reviewed sheet).

Applies the decisions Ken recorded on the exported V1 review-queue workbook
through the **existing V1 manual-review helpers** (``apply_manual_approval_state``
/ ``apply_manual_false_positive_state`` in ``app/api/alerts.py``) so publication
state + audit trail stay identical to the ``POST /api/v1/alerts/{id}/review`` path.
NEVER touches the legacy Jinja dashboard path and never writes raw SQL.

Decision vocabulary (normalized): ``approve_publish`` /
``publish_as_analyst_observation`` (both publish), ``reject_false_positive`` /
``reject_duplicate`` (both exclude). The backend has no Analyst-Observation content
type and no duplicate-specific ``pending_review_reason`` enum value, so:
  * Analyst Observation → published as a normal alert; the audit ``review_status``
    is marked ``analyst_observation`` and Ken's note is preserved in the report.
  * Duplicate → excluded via the false-positive helper; ``publish_decision_reason``
    + ``excluded_reason`` overridden to ``manual_duplicate`` (free-text), while
    ``pending_review_reason`` stays ``manual_rejected`` (the only valid enum).

Default is a read-only dry-run; a real apply requires the full guard set
(confirm token + a prior dry-run report + a matching file checksum + a fresh
dry-run drift check) and runs in a SINGLE transaction (rollback on any error).
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.alerts import apply_manual_approval_state, apply_manual_false_positive_state
from app.models.processed_alert import ProcessedAlert
from app.models.raw_item import RawItem
from app.models.review import AlertReview
from app.pipeline.publishing.constants import DecisionSource

CONFIRM_TOKEN = "APPLY_V1_REVIEW_QUEUE_KEN_20260622"
DEFAULT_BATCH_ID = "v1-review-queue-ken-2026-06-22"
DEFAULT_USER_ID = 1  # admin@hiddenalerts.com (role=admin)

# Canonical decisions for the 29 review-queue alerts (Ken, 2026-06-22). The parsed
# workbook must match this map exactly — it is the cross-check against the sheet.
EXPECTED: dict[int, str] = {
    **{i: "approve_publish" for i in (843, 838, 666, 76, 79, 888, 669, 645, 540, 490, 457, 92, 94)},
    **{i: "publish_as_analyst_observation" for i in (905, 833, 439)},
    **{i: "reject_false_positive" for i in (89, 816, 800, 794, 675, 556, 475, 437, 90, 91, 93)},
    **{i: "reject_duplicate" for i in (469, 458)},
}
EXPECTED_COUNTS = dict(Counter(EXPECTED.values()))  # {approve_publish:13, ...}

_PUBLISH_DECISIONS = {"approve_publish", "publish_as_analyst_observation"}
_AUDIT_STATUS = {
    "approve_publish": "approved",
    "publish_as_analyst_observation": "analyst_observation",
    "reject_false_positive": "false_positive",
    "reject_duplicate": "duplicate",
}

_ID_COLS = {"alert_id", "id"}
_DECISION_COLS = {"your review", "ken_decision", "decision", "review"}
_NOTES_COLS = {"your notes", "ken_notes", "notes"}


# ---------------------------------------------------------------------------
# Parsing + normalization
# ---------------------------------------------------------------------------


def normalize_decision(value) -> str | None:
    """Map a free-text decision cell to a normalized decision, or None if unknown."""
    if value is None:
        return None
    t = " ".join(str(value).split()).lower().replace("/", " ").replace("  ", " ").strip()
    if not t:
        return None
    if "duplicate" in t:
        return "reject_duplicate"
    if "analyst" in t or "observation" in t:
        return "publish_as_analyst_observation"
    if "false positive" in t or "reject" in t or "false-positive" in t:
        return "reject_false_positive"
    if "approve" in t or "publish" in t:
        return "approve_publish"
    return None


def _to_int(value):
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        try:
            f = float(str(value).strip())
        except ValueError:
            return None
        return int(f) if f.is_integer() else None


def file_checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def parse_workbook(path: Path) -> dict:
    """Read Ken's decisions from the 'Review Queue' sheet. Returns a parsed dict."""
    import openpyxl  # lazy

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "Review Queue" if "Review Queue" in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))
    headers = [(" ".join(str(c).split()) if c is not None else "") for c in (rows[0] if rows else [])]
    lower = [h.lower() for h in headers]

    def col(cands):
        for i, h in enumerate(lower):
            if h in cands:
                return i
        return None

    i_id, i_dec, i_note = col(_ID_COLS), col(_DECISION_COLS), col(_NOTES_COLS)

    out = {
        "sheet": sheet, "headers": headers,
        "id_col": i_id, "decision_col": i_dec, "notes_col": i_note,
        "decisions": {},          # alert_id -> normalized decision
        "raw": [],                # per-row raw record
        "notes": {},              # alert_id -> note (collapsed)
        "blank_decisions": [], "unknown_decisions": [], "blank_ids": [],
        "duplicate_ids": [], "non_integer_ids": [], "total_rows": 0,
    }
    if i_id is None or i_dec is None:
        out["fatal"] = f"missing_columns id_col={i_id} decision_col={i_dec} headers={headers}"
        return out

    seen: set[int] = set()
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out["total_rows"] += 1
        raw_id = r[i_id] if i_id < len(r) else None
        raw_dec = r[i_dec] if i_dec < len(r) else None
        note = r[i_note] if (i_note is not None and i_note < len(r)) else None
        aid = _to_int(raw_id)
        if aid is None:
            (out["blank_ids"] if raw_id in (None, "") else out["non_integer_ids"]).append(raw_id)
            continue
        norm = normalize_decision(raw_dec)
        out["raw"].append({"alert_id": aid, "raw_decision": raw_dec, "normalized": norm})
        if aid in seen:
            out["duplicate_ids"].append(aid)
            continue
        seen.add(aid)
        if norm is None:
            (out["blank_decisions"] if raw_dec in (None, "") else out["unknown_decisions"]).append(
                {"alert_id": aid, "raw_decision": raw_dec}
            )
            continue
        out["decisions"][aid] = norm
        out["notes"][aid] = " ".join(str(note).split()) if note not in (None, "") else ""
    return out


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_parsed(parsed: dict) -> list[str]:
    """Spreadsheet-level validation (no DB). Returns refusal reasons; empty = ok."""
    errors: list[str] = []
    if parsed.get("fatal"):
        return [parsed["fatal"]]
    if parsed["blank_ids"]:
        errors.append(f"blank_ids={len(parsed['blank_ids'])}")
    if parsed["non_integer_ids"]:
        errors.append(f"non_integer_ids={parsed['non_integer_ids']}")
    if parsed["duplicate_ids"]:
        errors.append(f"duplicate_ids={sorted(set(parsed['duplicate_ids']))}")
    if parsed["blank_decisions"]:
        errors.append(f"blank_decisions={[d['alert_id'] for d in parsed['blank_decisions']]}")
    if parsed["unknown_decisions"]:
        errors.append(f"unknown_decisions={parsed['unknown_decisions']}")

    decisions = parsed["decisions"]
    if len(decisions) != 29:
        errors.append(f"expected_29_rows_got={len(decisions)}")
    # Exact id-set + per-id decision match against the canonical EXPECTED map.
    got_ids, exp_ids = set(decisions), set(EXPECTED)
    if got_ids != exp_ids:
        errors.append(
            f"id_set_mismatch only_in_sheet={sorted(got_ids - exp_ids)} "
            f"only_in_expected={sorted(exp_ids - got_ids)}"
        )
    mismatched = {i: (decisions[i], EXPECTED[i]) for i in got_ids & exp_ids if decisions[i] != EXPECTED[i]}
    if mismatched:
        errors.append(f"decision_mismatch_vs_expected={mismatched}")
    counts = dict(Counter(decisions.values()))
    if counts != EXPECTED_COUNTS:
        errors.append(f"count_mismatch got={counts} expected={EXPECTED_COUNTS}")
    return errors


# ---------------------------------------------------------------------------
# Dry-run (read-only)
# ---------------------------------------------------------------------------


def _expected_change(decision: str) -> dict:
    if decision in _PUBLISH_DECISIONS:
        return {"is_published": True, "publish_decision": "auto_publish",
                "publish_decision_reason": "manual_admin_approved",
                "pending_review_reason": None, "is_excluded": False,
                "publication_state_source": "manual_admin",
                "audit_review_status": _AUDIT_STATUS[decision]}
    reason = "manual_duplicate" if decision == "reject_duplicate" else "manual_false_positive"
    return {"is_published": False, "publish_decision": "exclude",
            "publish_decision_reason": reason, "pending_review_reason": "manual_rejected",
            "is_excluded": True, "excluded_reason": reason,
            "publication_state_source": "manual_admin",
            "audit_review_status": _AUDIT_STATUS[decision]}


async def run_dry_run(input_path: str | Path, *, session: AsyncSession, batch_id: str) -> dict:
    path = Path(input_path)
    parsed = parse_workbook(path)
    errors = validate_parsed(parsed)

    decisions = parsed["decisions"]
    planned: list[dict] = []
    db_errors: list[str] = []
    missing, not_review, already_published, already_excluded = [], [], [], []

    if decisions:
        stmt = (
            select(ProcessedAlert)
            .where(ProcessedAlert.id.in_(list(decisions.keys())))
            .options(selectinload(ProcessedAlert.raw_item).selectinload(RawItem.source))
        )
        by_id = {a.id: a for a in (await session.execute(stmt)).scalars().all()}
        for aid in sorted(decisions):
            alert = by_id.get(aid)
            if alert is None:
                missing.append(aid)
                continue
            if alert.publish_decision != "review":
                not_review.append({"alert_id": aid, "current": alert.publish_decision})
            if alert.is_published:
                already_published.append(aid)
            if alert.is_excluded:
                already_excluded.append(aid)
            planned.append({
                "alert_id": aid,
                "decision": decisions[aid],
                "note": parsed["notes"].get(aid, ""),
                "current": {
                    "publish_decision": alert.publish_decision,
                    "pending_review_reason": alert.pending_review_reason,
                    "risk_band": alert.risk_band, "is_published": alert.is_published,
                    "is_excluded": alert.is_excluded, "is_relevant": alert.is_relevant,
                    "title": (alert.raw_item.title if alert.raw_item else None),
                },
                "would_change": _expected_change(decisions[aid]),
            })

    if missing:
        db_errors.append(f"alerts_not_found={missing}")
    if not_review:
        db_errors.append(f"not_in_review={not_review}")
    if already_published:
        db_errors.append(f"already_published={already_published}")
    if already_excluded:
        db_errors.append(f"already_excluded={already_excluded}")

    all_errors = errors + db_errors
    return {
        "mode": "dry_run",
        "batch_id": batch_id,
        "input_file": str(path),
        "input_sha256": file_checksum(path),
        "sheet": parsed.get("sheet"),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_rows_parsed": len(decisions),
        "counts_by_decision": dict(Counter(decisions.values())),
        "expected_counts": EXPECTED_COUNTS,
        "planned_actions": planned,
        "errors": all_errors,
        "passed": not all_errors,
        "db_changes_made": False,
    }


# ---------------------------------------------------------------------------
# Apply (guarded, single transaction)
# ---------------------------------------------------------------------------


def _refused(errors, **extra) -> dict:
    return {"mode": "apply_refused", "applied": False, "errors": errors, **extra}


def validate_for_apply(prior: dict, fresh: dict, *, input_sha256: str) -> list[str]:
    errors: list[str] = []
    if prior.get("mode") != "dry_run":
        errors.append(f"prior_mode_invalid={prior.get('mode')!r}")
    if prior.get("passed") is not True:
        errors.append("prior_report_not_passed")
    if prior.get("errors"):
        errors.append(f"prior_report_has_errors={prior.get('errors')}")
    if prior.get("input_sha256") != input_sha256:
        errors.append("input_checksum_changed_since_prior_report")
    if fresh.get("passed") is not True:
        errors.append(f"fresh_report_not_passed errors={fresh.get('errors')}")
    if fresh.get("input_sha256") != input_sha256:
        errors.append("fresh_input_checksum_mismatch")
    # plan agreement
    pa = {p["alert_id"]: p["decision"] for p in prior.get("planned_actions", [])}
    fa = {p["alert_id"]: p["decision"] for p in fresh.get("planned_actions", [])}
    if pa != fa:
        errors.append(f"plan_drift prior={pa} fresh={fa}")
    return errors


def _maybe_lock(stmt, session: AsyncSession):
    try:
        name = session.get_bind().dialect.name
    except Exception:
        name = ""
    return stmt.with_for_update() if name and name != "sqlite" else stmt


async def run_apply(
    input_path: str | Path,
    *,
    session: AsyncSession,
    apply: bool = False,
    dry_run_report_path: str | Path | None = None,
    batch_id: str | None = None,
    confirm: str | None = None,
    user_id: int = DEFAULT_USER_ID,
) -> dict:
    batch_id = batch_id or DEFAULT_BATCH_ID
    if not apply:
        report = await run_dry_run(input_path, session=session, batch_id=batch_id)
        report["applied"] = False
        return report

    if confirm != CONFIRM_TOKEN:
        return _refused([f"confirm_required: pass --confirm {CONFIRM_TOKEN}"])
    if not dry_run_report_path:
        return _refused(["dry_run_report_required_for_apply"])
    if not batch_id:
        return _refused(["batch_id_required_for_apply"])

    path = Path(input_path)
    try:
        prior = json.loads(Path(dry_run_report_path).read_text(encoding="utf-8"))
    except FileNotFoundError:
        return _refused([f"dry_run_report_not_found: {dry_run_report_path}"])
    except json.JSONDecodeError as exc:
        return _refused([f"dry_run_report_not_json: {exc}"])

    fresh = await run_dry_run(path, session=session, batch_id=batch_id)
    errors = validate_for_apply(prior, fresh, input_sha256=file_checksum(path))
    if errors:
        return _refused(errors, batch_id=batch_id, dry_run_report=str(dry_run_report_path))

    decisions = {p["alert_id"]: p["decision"] for p in fresh["planned_actions"]}
    notes = {p["alert_id"]: p.get("note", "") for p in fresh["planned_actions"]}
    ids = sorted(decisions)
    now = datetime.now(timezone.utc)
    applied = {d: [] for d in _AUDIT_STATUS}
    final_state: list[dict] = []
    reviews_created = 0

    try:
        stmt = (
            select(ProcessedAlert)
            .where(ProcessedAlert.id.in_(ids))
            .options(selectinload(ProcessedAlert.raw_item).selectinload(RawItem.source))
        )
        by_id = {a.id: a for a in (await session.execute(_maybe_lock(stmt, session))).scalars().all()}
        missing = [i for i in ids if i not in by_id]
        if missing:
            raise RuntimeError(f"alerts vanished between dry-run and apply: {missing}")
        drifted = [i for i in ids if by_id[i].publish_decision != "review"]
        if drifted:
            raise RuntimeError(f"alerts no longer in review: {drifted}")

        for aid in ids:
            alert = by_id[aid]
            decision = decisions[aid]
            if decision in _PUBLISH_DECISIONS:
                apply_manual_approval_state(alert, user_id, now=now)
            else:
                apply_manual_false_positive_state(alert, now=now)
                if decision == "reject_duplicate":
                    alert.publish_decision_reason = "manual_duplicate"
                    alert.excluded_reason = "manual_duplicate"
            session.add(AlertReview(
                alert_id=aid, user_id=user_id,
                review_status=_AUDIT_STATUS[decision],
                decision_source=DecisionSource.MANUAL_ADMIN.value,
                review_batch_id=batch_id,
            ))
            applied[decision].append(aid)
            reviews_created += 1
            final_state.append({
                "alert_id": aid, "decision": decision, "note": notes.get(aid, ""),
                "is_published": alert.is_published, "publish_decision": alert.publish_decision,
                "publish_decision_reason": alert.publish_decision_reason,
                "pending_review_reason": alert.pending_review_reason,
                "is_excluded": alert.is_excluded, "excluded_reason": alert.excluded_reason,
                "publication_state_source": alert.publication_state_source,
                "audit_review_status": _AUDIT_STATUS[decision],
            })

        await session.commit()
    except Exception as exc:  # noqa: BLE001
        await session.rollback()
        return _refused([f"apply_failed_rolled_back: {exc}"], batch_id=batch_id)

    return {
        "mode": "apply",
        "applied": True,
        "batch_id": batch_id,
        "input_file": str(path),
        "input_sha256": file_checksum(path),
        "generated_at": now.isoformat(),
        "counts": {
            "approve_publish": len(applied["approve_publish"]),
            "publish_as_analyst_observation": len(applied["publish_as_analyst_observation"]),
            "reject_false_positive": len(applied["reject_false_positive"]),
            "reject_duplicate": len(applied["reject_duplicate"]),
            "total_applied": reviews_created,
            "skipped": 0,
        },
        "applied_alert_ids": {d: sorted(applied[d]) for d in _AUDIT_STATUS},
        "review_records_created": reviews_created,
        "final_state": final_state,
        "publication_state_source": DecisionSource.MANUAL_ADMIN.value,
        "passed": True,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def console_summary(report: dict) -> str:
    mode = report.get("mode")
    if mode == "apply":
        c = report["counts"]
        return "\n".join([
            "V1 Review-Queue Decisions APPLY",
            f"Batch: {report['batch_id']}",
            f"Approved/published: {c['approve_publish']}  Analyst-observation: {c['publish_as_analyst_observation']}",
            f"False-positive: {c['reject_false_positive']}  Duplicate: {c['reject_duplicate']}",
            f"Total applied: {c['total_applied']}  Skipped: {c['skipped']}",
            f"Audit rows created: {report['review_records_created']}",
            f"Passed: {report['passed']}",
        ])
    if mode == "apply_refused":
        return "V1 Review-Queue APPLY REFUSED\n" + "\n".join(f"  - {e}" for e in report["errors"]) + \
            "\nNo database changes were applied."
    return "\n".join([
        "V1 Review-Queue Decisions Dry Run",
        f"Input: {report['input_file']}",
        f"sha256: {report['input_sha256']}",
        f"Rows parsed: {report['total_rows_parsed']}  counts={report['counts_by_decision']}",
        f"Errors: {report['errors']}",
        f"Passed: {report['passed']}",
        "No database changes were made.",
    ])


def write_markdown(report: dict, path: Path) -> None:
    lines = [f"# V1 Review-Queue Decisions — {report.get('mode')}", ""]
    lines.append(f"- Batch: `{report.get('batch_id')}`")
    lines.append(f"- Input: `{report.get('input_file')}`")
    lines.append(f"- Input sha256: `{report.get('input_sha256')}`")
    lines.append(f"- Generated: {report.get('generated_at')}")
    if report.get("mode") == "dry_run":
        lines.append(f"- Rows parsed: {report['total_rows_parsed']}  counts={report['counts_by_decision']}")
        lines.append(f"- Passed: {report['passed']}  errors={report['errors']}")
        lines.append("- **No database changes were made.**")
        lines += ["", "| alert_id | decision | current | → would_change | note |", "|---|---|---|---|---|"]
        for p in report["planned_actions"]:
            wc = p["would_change"]
            lines.append(
                f"| {p['alert_id']} | {p['decision']} | {p['current']['publish_decision']} | "
                f"{wc['publish_decision']} (pub={wc['is_published']}) | {(p['note'] or '')[:50].replace(chr(124),'/')} |"
            )
    elif report.get("mode") == "apply":
        c = report["counts"]
        lines.append(f"- Applied: approve={c['approve_publish']} analyst={c['publish_as_analyst_observation']} "
                     f"fp={c['reject_false_positive']} dup={c['reject_duplicate']} total={c['total_applied']}")
        lines.append(f"- Audit rows created: {report['review_records_created']}")
        lines.append("- **No deploy / restart / migration / backfill / candidate apply was run.**")
        lines += ["", "| alert_id | decision | published | publish_decision | reason | pending | audit |",
                  "|---|---|---|---|---|---|---|"]
        for s in report["final_state"]:
            lines.append(
                f"| {s['alert_id']} | {s['decision']} | {s['is_published']} | {s['publish_decision']} | "
                f"{s['publish_decision_reason']} | {s['pending_review_reason']} | {s['audit_review_status']} |"
            )
    else:
        lines.append(f"- REFUSED: {report.get('errors')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


async def _main_async(args: argparse.Namespace) -> int:
    from app.database import AsyncSessionLocal

    async with AsyncSessionLocal() as session:
        report = await run_apply(
            args.input, session=session, apply=args.apply,
            dry_run_report_path=args.dry_run_report, batch_id=args.batch_id,
            confirm=args.confirm, user_id=args.user_id,
        )

    out_dir = Path("reports")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    kind = {"apply": "apply_result", "apply_refused": "apply_refused"}.get(report.get("mode"), "apply_dry_run")
    json_path = out_dir / f"v1_review_queue_ken_{kind}_{ts}.json"
    md_path = out_dir / f"v1_review_queue_ken_{kind}_{ts}.md"
    json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    write_markdown(report, md_path)

    print(console_summary(report))
    print(f"\nJSON: {json_path}\nMD:   {md_path}")
    return 0 if report.get("passed") else 1


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Apply Ken's reviewed V1 review-queue decisions (dry-run by default).")
    p.add_argument("--input", required=True, help="Path to Ken's reviewed .xlsx workbook")
    p.add_argument("--apply", action="store_true", help="Apply (default: dry-run)")
    p.add_argument("--dry-run-report", help="Prior dry-run JSON report (required for --apply)")
    p.add_argument("--batch-id", default=DEFAULT_BATCH_ID)
    p.add_argument("--confirm", help=f"Must equal {CONFIRM_TOKEN} for --apply")
    p.add_argument("--user-id", type=int, default=DEFAULT_USER_ID)
    args = p.parse_args(argv)
    return asyncio.run(_main_async(args))


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
