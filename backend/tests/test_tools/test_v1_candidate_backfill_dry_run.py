"""V1 Slice 7B — Candidate Backfill dry-run tool tests.

Pure parse/normalize/count tests need no DB; DB-integration tests use the
in-memory session and a temporary CSV (never Ken's real spreadsheet). The tool
is read-only — one test proves zero mutation.
"""
from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

import pytest
from sqlalchemy import select

from app.models.processed_alert import ProcessedAlert
from app.models.raw_item import RawItem
from app.models.source import Source
from app.tools.v1_candidate_backfill_dry_run import (
    EXPECTED_COUNTS,
    detect_columns,
    normalize_decision,
    parse_records,
    read_input,
    run_dry_run,
    validate_counts,
)


def _write_csv(tmp_path: Path, rows: list[dict], headers=("alert_id", "decision")) -> Path:
    p = tmp_path / "decisions.csv"
    with open(p, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(headers))
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return p


# --- Pure parsing / normalization (no DB) ------------------------------------


def test_expected_counts_constant():
    assert EXPECTED_COUNTS == {"approved": 48, "rejected": 44, "keep_later": 27}


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("approved", "approved"), ("Approve", "approved"), ("YES", "approved"),
        ("controlled backfill", "approved"), ("publish", "approved"),
        ("rejected", "rejected"), ("Exclude", "rejected"), ("false positive", "rejected"),
        ("no", "rejected"),
        ("keep later", "keep_later"), ("manual hold", "keep_later"), ("Hold", "keep_later"),
        ("future review", "keep_later"),
        # Ken's reviewed-file exact decision values:
        ("Approve for controlled backfill", "approved"),
        ("Keep for later", "keep_later"),
        ("Reject", "rejected"),
        ("", None), (None, None), ("maybe", None),
    ],
)
def test_decision_aliases_normalize(raw, expected):
    assert normalize_decision(raw) == expected


@pytest.mark.parametrize(
    "headers,exp",
    [
        (["alert_id", "decision"], ("alert_id", "decision")),
        (["ID", "Ken Decision"], ("ID", "Ken Decision")),
        (["processed_alert_id", "status"], ("processed_alert_id", "status")),
        # Ken's reviewed-file header names:
        (["Alert ID", "Manual Review Decision"], ("Alert ID", "Manual Review Decision")),
        (["foo", "bar"], (None, None)),
    ],
)
def test_detect_columns(headers, exp):
    assert detect_columns(headers) == exp


def test_parse_counts_and_dedup():
    headers = ["alert_id", "decision"]
    rows = [
        {"alert_id": "1", "decision": "approved"},
        {"alert_id": "2", "decision": "reject"},
        {"alert_id": "3", "decision": "keep later"},
        {"alert_id": "3", "decision": "keep later"},  # duplicate same
    ]
    parsed = parse_records(headers, rows)
    assert parsed.counts == {"approved": 1, "rejected": 1, "keep_later": 1}
    assert parsed.duplicate_same == [3]
    assert parsed.duplicate_conflicting == []
    assert parsed.valid == {1: "approved", 2: "rejected", 3: "keep_later"}


def test_csv_spaced_headers_normalized(tmp_path):
    """Headers with surrounding whitespace must still be detected and row keys
    must resolve against the stripped header names."""
    p = tmp_path / "spaced.csv"
    p.write_text(
        " Alert ID , Manual Review Decision \n"
        "701,Approve for controlled backfill\n",
        encoding="utf-8",
    )
    headers, rows = read_input(p)
    assert detect_columns(headers) == ("Alert ID", "Manual Review Decision")
    parsed = parse_records(headers, rows)
    assert parsed.id_column == "Alert ID"
    assert parsed.decision_column == "Manual Review Decision"
    assert parsed.valid == {701: "approved"}
    assert parsed.counts["approved"] == 1
    assert parsed.invalid_rows == []
    assert parsed.invalid_decisions == []


def test_invalid_decision_reported():
    parsed = parse_records(
        ["alert_id", "decision"],
        [{"alert_id": "1", "decision": "approved"}, {"alert_id": "2", "decision": "huh"}],
    )
    assert len(parsed.invalid_decisions) == 1
    assert parsed.invalid_decisions[0]["alert_id"] == 2


def test_invalid_id_reported():
    parsed = parse_records(
        ["alert_id", "decision"],
        [{"alert_id": "abc", "decision": "approved"}, {"alert_id": "", "decision": "reject"}],
    )
    assert len(parsed.invalid_rows) == 2


def test_duplicate_conflicting_reported():
    parsed = parse_records(
        ["alert_id", "decision"],
        [{"alert_id": "5", "decision": "approved"}, {"alert_id": "5", "decision": "reject"}],
    )
    assert parsed.duplicate_conflicting == [5]
    assert 5 not in parsed.valid


def test_validate_counts_strict_default():
    res = validate_counts({"approved": 1, "rejected": 1, "keep_later": 1},
                          {"approved": 1, "rejected": 1, "keep_later": 2}, allow_mismatch=False)
    assert res["passed"] is False
    assert res["errors"]


def test_validate_counts_allow_mismatch():
    res = validate_counts({"approved": 1, "rejected": 1, "keep_later": 1},
                          {"approved": 1, "rejected": 1, "keep_later": 2}, allow_mismatch=True)
    assert res["passed"] is True
    assert res["strict_match"] is False


def test_xlsx_without_openpyxl_clear_error(tmp_path):
    """If openpyxl is absent, .xlsx read raises a clear error (not implemented here)."""
    p = tmp_path / "x.xlsx"
    p.write_bytes(b"not a real xlsx")
    try:
        import openpyxl  # noqa: F401
        has_openpyxl = True
    except ImportError:
        has_openpyxl = False
    if has_openpyxl:
        pytest.skip("openpyxl installed — clear-error path not applicable")
    with pytest.raises(RuntimeError, match="openpyxl"):
        read_input(p)


# --- DB integration (read-only) ----------------------------------------------


async def _seed_alert(db_session, *, idx, is_relevant=True, is_published=False, score=22):
    source = Source(
        name=f"Src {idx}", base_url=f"https://e{idx}.com", source_type="rss",
        credibility_score=4, adapter_class="RSSAdapter",
    )
    db_session.add(source)
    await db_session.flush()
    raw = RawItem(source_id=source.id, item_url=f"https://e{idx}.com/a", title=f"T{idx}",
                  url_hash=f"h{idx}")
    db_session.add(raw)
    await db_session.flush()
    alert = ProcessedAlert(
        raw_item_id=raw.id, primary_category="Cybercrime", risk_level="high",
        signal_score_total=score, is_relevant=is_relevant, is_published=is_published,
        matched_keywords=["fraud"], processed_at=datetime.now(timezone.utc),
        published_at=datetime.now(timezone.utc) if is_published else None,
    )
    db_session.add(alert)
    await db_session.commit()
    await db_session.refresh(alert)
    return alert


@pytest.mark.asyncio
async def test_passed_dry_run_small_expected(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=1)
    r = await _seed_alert(db_session, idx=2)
    k = await _seed_alert(db_session, idx=3)
    csv_path = _write_csv(tmp_path, [
        {"alert_id": a.id, "decision": "approved"},
        {"alert_id": r.id, "decision": "rejected"},
        {"alert_id": k.id, "decision": "keep later"},
    ])
    report = await run_dry_run(
        csv_path, session=db_session,
        expected_counts={"approved": 1, "rejected": 1, "keep_later": 1},
    )
    assert report["passed"] is True
    assert report["count_validation"]["passed"] is True
    assert report["db_validation"]["passed"] is True

    # Complete would_change previews (score 22 → risk_band "critical").
    assert report["planned_actions"]["approved"][0]["would_change"] == {
        "is_published": True,
        "publish_decision": "auto_publish",
        "publish_decision_reason": "candidate_backfill_approved",
        "pending_review_reason": None,
        "is_excluded": False,
        "excluded_reason": None,
        "is_manual_hold": False,
        "published_by_rule": False,
        "publication_state_source": "candidate_backfill",
        "publishing_policy_version": "v1.0",
        "risk_band": "critical",
    }
    assert report["planned_actions"]["rejected"][0]["would_change"] == {
        "is_published": False,
        "publish_decision": "exclude",
        "publish_decision_reason": "candidate_backfill_rejected",
        "pending_review_reason": "manual_rejected",
        "is_excluded": True,
        "excluded_reason": "candidate_backfill_rejected",
        "is_manual_hold": False,
        "published_by_rule": False,
        "publication_state_source": "candidate_backfill",
        "publishing_policy_version": "v1.0",
        "risk_band": "critical",
    }
    assert report["planned_actions"]["keep_later"][0]["would_change"] == {
        "is_published": False,
        "publish_decision": "hold",
        "publish_decision_reason": "candidate_backfill_keep_later",
        "pending_review_reason": "manual_hold",
        "is_excluded": False,
        "excluded_reason": None,
        "is_manual_hold": True,
        "published_by_rule": False,
        "publication_state_source": "candidate_backfill",
        "publishing_policy_version": "v1.0",
        "risk_band": "critical",
    }


@pytest.mark.asyncio
async def test_count_mismatch_fails_by_default(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=10)
    csv_path = _write_csv(tmp_path, [{"alert_id": a.id, "decision": "approved"}])
    report = await run_dry_run(csv_path, session=db_session)  # default 48/44/27
    assert report["passed"] is False
    assert "count_mismatch" in report["blockers"]


@pytest.mark.asyncio
async def test_count_mismatch_allowed_flag(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=11)
    csv_path = _write_csv(tmp_path, [{"alert_id": a.id, "decision": "approved"}])
    report = await run_dry_run(csv_path, session=db_session, allow_count_mismatch=True)
    assert "count_mismatch" not in report["blockers"]
    assert report["count_validation"]["passed"] is True


@pytest.mark.asyncio
async def test_missing_db_id_reported(db_session, tmp_path):
    csv_path = _write_csv(tmp_path, [{"alert_id": 999999, "decision": "approved"}])
    report = await run_dry_run(
        csv_path, session=db_session, expected_counts={"approved": 1, "rejected": 0, "keep_later": 0},
    )
    assert 999999 in report["db_validation"]["missing_ids"]
    assert report["passed"] is False


@pytest.mark.asyncio
async def test_approved_already_published_blocker(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=20, is_published=True)
    csv_path = _write_csv(tmp_path, [{"alert_id": a.id, "decision": "approved"}])
    report = await run_dry_run(
        csv_path, session=db_session, expected_counts={"approved": 1, "rejected": 0, "keep_later": 0},
    )
    assert a.id in report["db_validation"]["approved_already_published_ids"]
    assert report["passed"] is False


@pytest.mark.asyncio
async def test_approved_irrelevant_blocker(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=30, is_relevant=False)
    csv_path = _write_csv(tmp_path, [{"alert_id": a.id, "decision": "approved"}])
    report = await run_dry_run(
        csv_path, session=db_session, expected_counts={"approved": 1, "rejected": 0, "keep_later": 0},
    )
    assert a.id in report["db_validation"]["approved_irrelevant_ids"]
    assert report["passed"] is False


@pytest.mark.asyncio
async def test_dry_run_does_not_mutate_db(db_session, tmp_path):
    a = await _seed_alert(db_session, idx=40, is_published=True, is_relevant=True)
    before = {
        "is_published": a.is_published, "is_relevant": a.is_relevant,
        "risk_band": a.risk_band, "publish_decision": a.publish_decision,
        "is_excluded": a.is_excluded, "is_manual_hold": a.is_manual_hold,
        "publication_state_source": a.publication_state_source,
        "signal_score_total": a.signal_score_total,
    }
    # Decisions that, if applied, WOULD change all of these.
    csv_path = _write_csv(tmp_path, [{"alert_id": a.id, "decision": "rejected"}])
    await run_dry_run(
        csv_path, session=db_session, expected_counts={"approved": 0, "rejected": 1, "keep_later": 0},
    )
    refreshed = (
        await db_session.execute(select(ProcessedAlert).where(ProcessedAlert.id == a.id))
    ).scalar_one()
    await db_session.refresh(refreshed)
    assert refreshed.is_published == before["is_published"]
    assert refreshed.is_relevant == before["is_relevant"]
    assert refreshed.risk_band == before["risk_band"]
    assert refreshed.publish_decision == before["publish_decision"]
    assert refreshed.is_excluded == before["is_excluded"]
    assert refreshed.is_manual_hold == before["is_manual_hold"]
    assert refreshed.publication_state_source == before["publication_state_source"]
    assert refreshed.signal_score_total == before["signal_score_total"]


@pytest.mark.asyncio
async def test_xlsx_parsing_if_openpyxl_available(db_session, tmp_path):
    """If openpyxl is present, .xlsx parses identically to CSV."""
    openpyxl = pytest.importorskip("openpyxl")
    a = await _seed_alert(db_session, idx=50)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["alert_id", "decision"])
    ws.append([a.id, "approved"])
    xlsx_path = tmp_path / "decisions.xlsx"
    wb.save(xlsx_path)
    report = await run_dry_run(
        xlsx_path, session=db_session, expected_counts={"approved": 1, "rejected": 0, "keep_later": 0},
    )
    assert report["detected_columns"] == {"id": "alert_id", "decision": "decision"}
    assert report["actual_counts"]["approved"] == 1
    assert report["passed"] is True


@pytest.mark.asyncio
async def test_xlsx_multisheet_autodetects_data_sheet(db_session, tmp_path):
    """A workbook with Read Me / Summary / Review Data sheets (like Ken's file)
    auto-selects the sheet exposing the id+decision columns."""
    openpyxl = pytest.importorskip("openpyxl")
    a = await _seed_alert(db_session, idx=60)
    wb = openpyxl.Workbook()
    wb.active.title = "Read Me"
    wb.active.append(["HiddenAlerts V1 Candidate Alerts Review"])  # banner, no columns
    summary = wb.create_sheet("Summary")
    summary.append(["Total rows", 1])
    data = wb.create_sheet("Review Data")
    data.append(["Alert ID", "Manual Review Decision"])
    data.append([a.id, "Approve for controlled backfill"])
    xlsx_path = tmp_path / "multi.xlsx"
    wb.save(xlsx_path)
    report = await run_dry_run(
        xlsx_path, session=db_session, expected_counts={"approved": 1, "rejected": 0, "keep_later": 0},
    )
    assert report["detected_columns"] == {"id": "Alert ID", "decision": "Manual Review Decision"}
    assert report["actual_counts"]["approved"] == 1
    assert report["passed"] is True
